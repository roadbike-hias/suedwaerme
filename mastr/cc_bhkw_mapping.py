import os
import json
import pandas as pd
import requests
from shapely.geometry import shape, Point, mapping
from datetime import datetime
from typing import List, Dict, Any, Tuple
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import Hyperlink

# Constants
min_years_old = 5  # Set to 0 to disable age filtering
GEOJSON_FOLDER = os.path.join(os.path.dirname(__file__), "..", "gebieteCC")
CSV_FILE = os.path.join(os.path.dirname(__file__), "Stromerzeuger.csv")
MATCHED_CSV = os.path.join(os.path.dirname(__file__), "matched_units.csv")
UNMATCHED_CSV = os.path.join(os.path.dirname(__file__), "unmatched_units.csv")
MATCHED_EXCEL = os.path.join(os.path.dirname(__file__), "matched_units.xlsx")
UNMATCHED_EXCEL = os.path.join(os.path.dirname(__file__), "unmatched_units.xlsx")
MATCHED_GEOJSON = os.path.join(os.path.dirname(__file__), "matched_units.geojson")
UNMATCHED_GEOJSON = os.path.join(os.path.dirname(__file__), "unmatched_units.geojson")

# Column names mapping
COLUMNS = {
    'latitude': "Koordinate: Breitengrad (WGS84)",
    'longitude': "Koordinate: Längengrad (WGS84)",
    'start_date': "Inbetriebnahmedatum der Einheit",
    'unit_name': "Anzeige-Name der Einheit",
    'power': "Bruttoleistung der Einheit",
    'zip_code': "Postleitzahl",
    'city': "Ort",
    'street': "Straße",
    'street_number': "Hausnummer",
    'operator': "Name des Anlagenbetreibers (nur Org.)",
    'mastr_number': "MaStR-Nr. der Einheit"
}

# Headers for MaStR request
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:135.0) Gecko/20100101 Firefox/135.0",
    "Accept": "*/*",
    "Accept-Language": "de,en-US;q=0.7,en;q=0.3",
    "Accept-Encoding": "gzip, deflate, br, zstd",
    "Origin": "https://www.eex.com",
    "DNT": "1",
    "Sec-GPC": "1",
    "Connection": "keep-alive",
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "no-cors",
    "Sec-Fetch-Site": "cross-site",
    "Priority": "u=0",
    "Pragma": "no-cache",
    "Cache-Control": "no-cache",
    "Referer": "https://www.eex.com/"
}

def load_geojson_files(folder_path: str) -> List[Dict[str, Any]]:
    """Load all GeoJSON files from a folder and return their features."""
    all_features = []
    for filename in os.listdir(folder_path):
        if filename.endswith(".geojson"):
            file_path = os.path.join(folder_path, filename)
            with open(file_path, "r", encoding='utf-8') as f:
                geojson_data = json.load(f)
                if geojson_data["type"] == "FeatureCollection":
                    all_features.extend(geojson_data["features"])
                elif geojson_data["type"] == "Feature":
                    all_features.append(geojson_data)
    return all_features


def parse_date(date_str: str) -> datetime:
    """Parse date string in format dd.mm.yyyy to datetime object."""
    try:
        return datetime.strptime(date_str, '%d.%m.%Y')
    except ValueError:
        return None


def calculate_age(start_date: str) -> int:
    """Calculate age in years from start date string."""
    parsed_date = parse_date(start_date)
    if not parsed_date:
        return None
    return relativedelta(datetime.now(), parsed_date).years


def filter_by_age(df: pd.DataFrame, min_years: int) -> pd.DataFrame:
    """Filter DataFrame to only include plants older than min_years."""
    if min_years <= 0:
        return df

    today = datetime.now()
    cutoff_date = today - relativedelta(years=min_years)

    df['parsed_date'] = df[COLUMNS['start_date']].apply(parse_date)
    filtered_df = df[df['parsed_date'] <= cutoff_date].copy()
    filtered_df.drop(columns=['parsed_date'], inplace=True)

    print(f"Filtered to {len(filtered_df)} plants older than {min_years} years")
    return filtered_df


def create_geojson(features: List[Dict]) -> Dict:
    """Create a GeoJSON FeatureCollection from features."""
    return {
        "type": "FeatureCollection",
        "features": features
    }


def get_final_mastr_url(mastr_number: str) -> str:
    """Generate MaStR URL from MaStR number and follow redirect to get final URL."""
    if not mastr_number or len(mastr_number) <= 3:
        return ""

    stripped_number = mastr_number[3:]
    initial_url = f"https://www.marktstammdatenregister.de/MaStR/Schnellsuche/Schnellsuche"

    # Parameters for MaStR request
    params = {
        "praefix": 'SEE',
        "mastrNummerOrId": stripped_number
    }

    try:
        response = requests.get(initial_url, headers=headers, params=params)

        if response.status_code == 200:
            return f"https://www.marktstammdatenregister.de{response.json()['url']}"
        return initial_url
    except requests.RequestException:
        return initial_url


def match_points_to_polygons(points: List[Point], features: List[Dict[str, Any]], df: pd.DataFrame) -> Tuple[
    pd.DataFrame, pd.DataFrame, Dict, Dict]:
    """Match points to polygons and return matched and unmatched DataFrames and GeoJSONs."""
    matched_features = []
    unmatched_features = []
    matched_data = []
    unmatched_data = []

    for idx, point in enumerate(points):
        row = df.iloc[idx]
        age = calculate_age(row[COLUMNS['start_date']])
        location = f"{row[COLUMNS['street']]} {row[COLUMNS['street_number']]}, {row[COLUMNS['zip_code']]} {row[COLUMNS['city']]}"
        mastr_url = get_final_mastr_url(row[COLUMNS['mastr_number']])

        properties = {
            'power': int(row[COLUMNS['power']]),
            'operator': row[COLUMNS['operator']],
            'unit_name': row[COLUMNS['unit_name']],
            'start_date': row[COLUMNS['start_date']],
            'age_years': age,
            'location': location,
            'mastr_number': str(row[COLUMNS['mastr_number']]),
            'mastr_url': str(mastr_url)
        }

        feature = {
            "type": "Feature",
            "geometry": mapping(point),
            "properties": properties
        }

        matched = False
        for poly_feature in features:
            polygon = shape(poly_feature['geometry'])
            if polygon.contains(point):
                properties['cc'] = poly_feature['properties'].get('CC', 'Unknown')
                matched_features.append(feature)
                matched_data.append(
                    {**properties, 'latitude': row[COLUMNS['latitude']], 'longitude': row[COLUMNS['longitude']]})
                matched = True
                break

        if not matched:
            properties['cc'] = 'No matching CC area'
            unmatched_features.append(feature)
            unmatched_data.append(
                {**properties, 'latitude': row[COLUMNS['latitude']], 'longitude': row[COLUMNS['longitude']]})

    matched_geojson = create_geojson(matched_features)
    unmatched_geojson = create_geojson(unmatched_features)

    return (
        pd.DataFrame(matched_data),
        pd.DataFrame(unmatched_data),
        matched_geojson,
        unmatched_geojson
    )


def save_geojson(data: Dict, filepath: str):
    """Save GeoJSON data to file."""
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def save_to_excel(df: pd.DataFrame, filepath: str):
    """Save DataFrame to Excel with proper date formatting and hyperlinks."""
    # Create a new workbook
    wb = Workbook()
    ws = wb.active

    # Write the data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            if r_idx == 1:  # Header row
                ws.cell(row=r_idx, column=c_idx, value=value)
            else:
                col_name = df.columns[c_idx - 1]
                if col_name == 'start_date':
                    # Format as date
                    try:
                        date_val = datetime.strptime(value, '%d.%m.%Y')
                        cell = ws.cell(row=r_idx, column=c_idx, value=date_val)
                        cell.number_format = 'dd.mm.yyy'
                    except (ValueError, TypeError):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                elif col_name == 'mastr_number' and 'mastr_url' in df.columns:
                    # Create hyperlink
                    url = df.iloc[r_idx - 2]['mastr_url']
                    if pd.notna(url) and url:
                        cell = ws.cell(row=r_idx, column=c_idx)
                        cell.value = value
                        cell.hyperlink = Hyperlink(ref=cell.coordinate, target=url, display=value)
                        cell.style = "Hyperlink"
                    else:
                        ws.cell(row=r_idx, column=c_idx, value=value)
                else:
                    ws.cell(row=r_idx, column=c_idx, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filepath)


def main():
    # Load and prepare data
    df = pd.read_csv(CSV_FILE, delimiter=";", decimal=",", encoding='utf-8', dtype={'Hausnummer': str})
    print(f"Loaded {len(df)} plants from CSV")

    # Filter by age (change the number to your desired minimum years)
    df = filter_by_age(df, min_years_old)

    features = load_geojson_files(GEOJSON_FOLDER)
    print(f"Loaded {len(features)} GeoJSON features")

    # Create points and match to polygons
    points = []
    for _, row in df.iterrows():
        try:
            lat = float(row[COLUMNS['latitude']])
            lon = float(row[COLUMNS['longitude']])
            points.append(Point(lon, lat))
        except (ValueError, KeyError) as e:
            print(f"Error processing row {_}: {e}")

    matched_df, unmatched_df, matched_geojson, unmatched_geojson = match_points_to_polygons(points, features, df)

    # Save CSV results
    matched_df.to_csv(MATCHED_CSV, index=False, encoding='utf-8')
    unmatched_df.to_csv(UNMATCHED_CSV, index=False, encoding='utf-8')

    # Save Excel results with formatting
    save_to_excel(matched_df, MATCHED_EXCEL)
    save_to_excel(unmatched_df, UNMATCHED_EXCEL)

    # Save GeoJSON results
    save_geojson(matched_geojson, MATCHED_GEOJSON)
    save_geojson(unmatched_geojson, UNMATCHED_GEOJSON)

    print("\nResults:")
    print(f"- Matched plants: {len(matched_df)}")
    print(f"  - CSV: {MATCHED_CSV}")
    print(f"  - Excel: {MATCHED_EXCEL} (with formatted dates and hyperlinks)")
    print(f"  - GeoJSON: {MATCHED_GEOJSON}")
    print(f"- Unmatched plants: {len(unmatched_df)}")
    print(f"  - CSV: {UNMATCHED_CSV}")
    print(f"  - Excel: {UNMATCHED_EXCEL} (with formatted dates and hyperlinks)")
    print(f"  - GeoJSON: {UNMATCHED_GEOJSON}")

    if len(matched_df) > 0:
        print("\nMatched plants by CC area:")
        print(matched_df['cc'].value_counts())


if __name__ == "__main__":
    main()